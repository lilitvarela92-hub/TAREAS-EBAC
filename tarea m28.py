# %% [markdown]
# ## **TAREA M28 - Liliana Tamalatzi Varela**

# %% [markdown]
# #### **Objetivo:** Cubrir las interacciones básicas con diferentes tipos de archivos: CSV, JSON y Excel, realizar operaciones básicas de lectura y escritura sobre los mismos y revisar las estructuras internas para el uso de datos, como listas, diccionarios y listas anidadas.

# %% [markdown]
# #### **Paso a paso:**
# Generar un archivo tipo Notebook que contenga el código fuente de varios ejercicios aplicados a tres fuentes de datos:
# Para cada archivo:
# - Leer el archivo y guardarlo en una lista
# - Obtener las primeras 10 filas
# - Obtener las ultimas 10 filas
# - Seleccionar 2-4 columnas del dataset y generar un archivo nuevo
# 

# %%
# Importar función os para obtener la dirección actual del trabajo 
import os
cwd = os.getcwd()
print("Directorio de Trabajo: {0}".format(cwd))

# %%
# Cambiar la dirección actual con os.chidr
os.chdir('C:/Users/LILO/Downloads')

# %% [markdown]
# ##
# ----------------------
# **ARCHIVO CSV**
# ##
# ----------------------

# %%
# Importar librerías
import csv
import numpy as np

# --- 1. Leer el archivo y guardarlo en una lista
archivo_csv = "fifa_eda.csv"
datos = []

with open(archivo_csv, newline='', encoding='utf-8') as archivo_csv:
    lector = csv.reader(archivo_csv)
    encabezado = next(lector)  # Leer la primera fila (nombres de columnas)
    for fila in lector:
        datos.append(fila)

encabezado

# %%
# --- 2. Imprime las 10 primeras filas
print("\n Primeras 10 filas:")
for fila in datos[:10]:
    print(fila)

# --- 3. Imprime las 10 últimas filas
print("\n Últimas 10 filas:")
for fila in datos[-10:]:
    print(fila)

# %%

# --- 4. Seleccionar nuevas columnas del Data Set y guardarlos en un nuevo archivo
columnas_seleccionadas = ['ID', 'Name', 'Age', 'Nationality']
datos_filtrados = []

indices = [encabezado.index(c) for c in columnas_seleccionadas]

datos_filtrados = [[fila[i] for i in indices] for fila in datos]

with open ('fifa_new.csv', 'w', encoding= 'UTF8', newline='') as f:
    writer = csv.writer(f)

    # Wirite the header 
    writer.writerow(columnas_seleccionadas)

    # Wirite the data
    writer.writerows(datos_filtrados)

# %%
# Luego lo podemos visualizar

# Abre el archivo de lectura
file = open(r'./fifa_new.csv')
# Lee el archivo con el objeto csvreader
csvreader = csv.reader(file)

# Obtiene los encabezados del archivo (header)
header_fifa = []
header_fifa = next(csvreader)
header_fifa

# %% [markdown]
# ###
# ------------------
# **ARCHIVO XLSX**
# ###
# ------------------

# %%
# Importar librerías
import openpyxl 
from pathlib import Path

xlsx_file = Path('.', 'Amazon_top100_bestselling_books_2009to2021.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file)

# ---- 1. Leer el archivo y guardarlo en una lista 
sheet = wb_obj.active

col_names = []
for column in sheet.iter_cols(1, sheet.max_column):
    col_names.append(column[0].value)

print(col_names)

# %%
# ---- 2. Obtener las primeras 10 filas
total_filas = sheet.max_row
print("Primeras 10 filas: \n")
for row in sheet.iter_rows (min_row= 1, max_col= 10, max_row= 12, values_only= True):
    print(row)

# %%
# ---- 3. Obtener las últimas 10 filas
print("Últimas 10 filas: \n")
for row in sheet.iter_rows(min_row= total_filas - 9, max_col= 10, max_row= total_filas, values_only= True):
    print(row)

# %%
# ---- 4. Seleccionar columnas y generar un nuevo archivo 
import xlsxwriter

datos_filtrados = []
for fila in sheet.iter_rows(values_only=True):
    nueva_fila = [fila[i] for i in indices]
    datos_filtrados.append(nueva_fila)

# %%
nuevo_archivo = Path('.', 'Amazon_books_filtrado.xlsx')
workbook = xlsxwriter.Workbook(nuevo_archivo)
hoja_nueva = workbook.add_worksheet('Filtrado')

# %%
# Escribir los datos en el nuevo archivo
for fila_idx, fila in enumerate(datos_filtrados):
    for col_idx, valor in enumerate(fila):
        hoja_nueva.write(fila_idx, col_idx, valor)

workbook.close()

print("Archivo creado correctamente.")

# %% [markdown]
# ###
# ------------------
# **ARCHIVO JSON**
# ###
# ------------------

# %%
# Importar librerias
import json 
import zipfile

# Ruta del archivo ZIP original
zip_path = "GOT Subtitles.zip"

# Almacenar las opciones reducidas
reducidos = {}

# %%
# Abrir el ZIP sin descomprimirlo
with zipfile.ZipFile(zip_path, 'r') as z:
    # Listar archivos dentro del ZIP
    print("Archivos encontrados en el ZIP:")
    print(z.namelist())
    print("-" * 60)

    for file_name in z.namelist():
        if file_name.endswith(".json"):
            print(f"\nProcesando: {file_name}")

            # Leer el contenido directamente (bytes → str)
            contenido = z.read(file_name).decode("utf-8")

            # Convertir a diccionario
            data = json.loads(contenido)

            # Convertir a lista de registros (temporada, episodio, línea, texto)
            registros = []
            temporada = file_name.split(".")[0]
            for episodio, dialogos in data.items():
                for linea, texto in dialogos.items():
                    registros.append({
                        "temporada": temporada,
                        "episodio": episodio,
                        "linea": int(linea),
                        "texto": texto
                    })
            # Mostrar primeras y últimas 10 filas
            print("\nPrimeras 10 filas:")
            for r in registros[:10]:
                print(r)
            print("\nÚltimas 10 filas:")
            for r in registros[-10:]:
                print(r)
            
             # Guardar el contenido procesado en memoria (no escribir aún)
            reducidos[file_name.replace(".json", "_reducido.json")] = registros

# %%
# --- CREACIÓN DEL NUEVO ZIP ---
with zipfile.ZipFile("GOT_reducido.zip", "w") as new_zip:
    for nombre_archivo, registros in reducidos.items():
        json_data = json.dumps(registros, indent=2, ensure_ascii=False)
        new_zip.writestr(nombre_archivo, json_data)

print("\n Proceso completado. Se creó el archivo 'GOT_reducido.zip'.")

# %%



